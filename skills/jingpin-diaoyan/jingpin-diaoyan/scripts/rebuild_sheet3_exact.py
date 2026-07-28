# -*- coding: utf-8 -*-
import shutil, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

src = r'C:\Users\胡康杰\Desktop\竞品调研框架-v23.xlsx'
out = r'C:\Users\胡康杰\Desktop\竞品调研框架-v24.xlsx'

shutil.copy2(src, out)
wb = openpyxl.load_workbook(out)
ws = wb['竞品产品对比']

# ===== 1. 精确颜色 =====
# 原始Sheet4颜色：
# A列：FF6699CC
# R2品牌列：FFFFFF00（黄色）
# C14(童状元)品牌列：FFF5B9D2（粉色）
# C15(龙承匠人)品牌列：00000000（无色=白色）
# 正文单元格：FFFFFFFF（白色）

a_fill = PatternFill(start_color='FF6699CC', end_color='FF6699CC', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
pink_fill = PatternFill(start_color='FFF5B9D2', end_color='FFF5B9D2', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
no_fill = PatternFill(fill_type=None)
transparent_fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')

a_font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
r1_link_font = Font(name='微软雅黑', size=8, color='2563EB', underline='single', bold=False)
white_font_bold = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
data_font = Font(name='微软雅黑', size=9, bold=False)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
no_wrap_center = Alignment(horizontal='center', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# ===== 2. 列宽修正 =====
ws.column_dimensions['A'].width = 11.6083333333333
ws.column_dimensions['B'].width = 13.0
ws.column_dimensions['C'].width = 13.0
ws.column_dimensions['D'].width = 11.25
ws.column_dimensions['E'].width = 13.0
ws.column_dimensions['F'].width = 13.0
ws.column_dimensions['G'].width = 13.0
ws.column_dimensions['H'].width = 13.0
ws.column_dimensions['I'].width = 13.0
ws.column_dimensions['J'].width = 14.9083333333333
ws.column_dimensions['K'].width = 13.0
ws.column_dimensions['L'].width = 13.0
ws.column_dimensions['M'].width = 14.2833333333333
ws.column_dimensions['N'].width = 13.0
ws.column_dimensions['O'].width = 19.0166666666667

# ===== 3. 行高修正 =====
row_heights = {
    1: 67.5, 2: 14.25, 3: 42.75, 4: 42.5,
    5: 28.5, 6: 28.5, 7: 28.5, 8: 28.5,
    9: 42.75, 10: 28.5, 11: 14.25, 12: 28.5,
    13: 14.25, 14: 28.5, 15: 14.25, 16: 14.25,
    17: 42.75, 18: 28.5, 19: 28.5, 20: 28.5,
    21: 14.25, 22: 14.25, 23: 14.25, 24: 28.5,
    25: 14.25, 26: 14.25, 27: 14.25, 28: 142.5,
}
for r, h in row_heights.items():
    ws.row_dimensions[r].height = h

# ===== 4. A列样式（R1~R28）= ####
# 原始Sheet4中A列所有单元格：6699CC蓝底、白字、加粗、左对齐
# R1与其余行一致
for r in range(1, 29):
    cell = ws.cell(r, 1)
    cell.fill = a_fill
    cell.font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
    cell.alignment = left_center
    cell.border = thin_border

# A列保留已有内容（参数维度名称）
a_labels = {
    1: '商品链接 >>>', 2: '品牌', 3: '产品名', 4: '图片',
    5: '活动价格(基础款)', 6: '主材', 7: '环保等级', 8: '尺寸',
    9: '详细尺寸(1.2M为例)', 10: '功能描述', 11: '桌面倾斜', 12: '安全配置',
    13: '追背/护腰', 14: '椅轮类型', 15: '升降方式', 16: '升降高度',
    17: '书架', 18: '储物配件', 19: '可成长配件', 20: '护眼灯',
    21: '连接件材质', 22: '涂装工艺', 23: '发货周期', 24: '配套椅子',
    25: '配套椅价格', 26: '儿童友好度', 27: '综合推荐指数', 28: '推荐理由'
}
for r, label in a_labels.items():
    ws.cell(r, 1).value = label

# ===== 5. R1链接行 =====
# 原始：白底、蓝色8号字、下划线、左对齐
for c in range(2, 16):
    cell = ws.cell(1, c)
    cell.fill = white_fill
    cell.font = r1_link_font
    cell.alignment = left_center
    cell.border = thin_border

# ===== 6. R2品牌行 =====
# 原始：C2~C13黄底、C14粉色底(FFF5B9D2)、C15无底(00000000=白色)
# 所有白字9号居中
r2_fills = {}
for c in range(2, 14):
    r2_fills[c] = yellow_fill
r2_fills[14] = pink_fill  # N=枝乐(对照原始C14=童状元)
r2_fills[15] = transparent_fill  # O=源氏木语(对照原始C15=龙承匠人)

for c in range(2, 16):
    cell = ws.cell(2, c)
    cell.fill = r2_fills.get(c, white_fill)
    cell.font = Font(name='微软雅黑', size=9, bold=True, color='000000')
    cell.alignment = data_align
    cell.border = thin_border

# ===== 7. R3产品名行 =====
# 原始：白底、9号字、居中
for c in range(2, 16):
    cell = ws.cell(3, c)
    cell.fill = white_fill
    cell.font = data_font
    cell.alignment = data_align
    cell.border = thin_border

# ===== 8. R4图片行 =====
# 原始：白底42.5行高，图片自己定位
for c in range(2, 16):
    cell = ws.cell(4, c)
    cell.fill = white_fill
    cell.font = data_font
    cell.alignment = data_align
    cell.border = thin_border

# ===== 9. R5~R28数据行 =====
# 原始：白底、9号字、居中
for r in range(5, 29):
    for c in range(2, 16):
        cell = ws.cell(r, c)
        cell.fill = white_fill
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border

# ===== 10. 重新嵌入图片（图片放R4，行高42.5） =====
# 图片不能太大，否则会超出，保持100x100
img_dir = r'C:\Users\胡康杰\Desktop\竞品调研脚本\product_imgs'
ws._images = []

img_map = {
    '无名字学习桌': None,
    '黑白调s2儿童学习桌': '黑白调s2儿童学习桌',
    '黑白调c2儿童学习桌': '黑白调c2儿童学习桌',
    '黑白调A2儿童学习桌': '黑白调A2儿童学习桌',
    '爱果乐沉浸岛学习桌': '爱果乐沉浸岛学习桌',
    '爱果乐学习桌咖啡猫': '爱果乐学习桌咖啡猫',
    '爱果乐木木岛': '爱果乐木木岛',
    '爱果乐艺简': '爱果乐艺简',
    '护童学习桌': '护童学习桌',
    '青节学习桌': '青节学习桌',
    '宜家学习桌': '宜家学习桌',
    '龙承匠人学习桌': '龙承匠人学习桌',
    '枝乐学习桌': '枝乐学习桌',
    '源氏木语学习桌': '源氏木语学习桌',
}

for c in range(2, 16):
    cl = get_column_letter(c)
    prod = ws.cell(3, c).value or ''
    img_key = img_map.get(prod)
    if img_key:
        for f in os.listdir(img_dir):
            if f.endswith('.jpg') and os.path.getsize(os.path.join(img_dir,f)) > 2000:
                base = f.replace('.jpg', '')
                if base == img_key or base.startswith(img_key):
                    try:
                        img = XlImage(os.path.join(img_dir, f))
                        img.width = 100
                        img.height = 100
                        img.anchor = f'{cl}4'
                        ws.add_image(img)
                        print(f'  ✅ 图 {cl}4: {prod}')
                    except Exception as e:
                        print(f'  ❌ {cl}: {e}')
                    break

wb.save(out)
wb.close()

print(f'\n✅ 保存至: {out}')

# 核查
wb2 = openpyxl.load_workbook(out)
ws2 = wb2['竞品产品对比']
print(f'A列R1填色: {ws2.cell(1,1).fill.start_color.rgb}')
print(f'R2C2填色: {ws2.cell(2,2).fill.start_color.rgb}')
print(f'R2C14填色: {ws2.cell(2,14).fill.start_color.rgb}')
print(f'M列宽: {ws2.column_dimensions["M"].width}')
print(f'R4行高: {ws2.row_dimensions[4].height}')
wb2.close()
