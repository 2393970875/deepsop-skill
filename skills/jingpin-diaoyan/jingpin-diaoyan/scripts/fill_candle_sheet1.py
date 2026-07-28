# -*- coding: utf-8 -*-
"""Phase 2: Sheet1 SEM关键词填入"""
import csv, openpyxl, os, io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DESKTOP = r"C:\Users\胡康杰\Desktop"
DST = os.path.join(DESKTOP, "竞品调研框架-香薰蜡烛v2.xlsx")
CSV_PATH = r"C:\Users\胡康杰\.openclaw\media\outbound\99222dab-186f-4b1b-b74e-2353a9d4a172.csv"

shutil = __import__('shutil')
shutil.copy2(DST.replace("v2", ""), DST)

# 读取CSV
rows = []
with io.open(CSV_PATH, encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

print(f"读到 {len(rows)} 条关键词")

# 打开Excel
wb = openpyxl.load_workbook(DST)
ws1 = wb["1.竞争品牌定位"]

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
data_font = Font(name='Arial', size=10)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# R3已有表头，从R4开始填入
# 把159条分成3组
group_count = len(rows) // 3
group1 = rows[:group_count]  # 第1组约53条
group2 = rows[group_count:2*group_count]
group3 = rows[2*group_count:]

groups = [group1, group2, group3]

for g_idx, group in enumerate(groups):
    base_col = g_idx * 5
    for i, r in enumerate(group):
        row = 4 + i  # 从R4开始
        keyword = r.get('\ufeff关键词', r.get('关键词', ''))
        ws1.cell(row, base_col + 1, keyword).font = data_font
        ws1.cell(row, base_col + 1).alignment = data_align
        ws1.cell(row, base_col + 1).border = thin_border

        ws1.cell(row, base_col + 2, r.get('推荐理由', '')).font = data_font
        ws1.cell(row, base_col + 2).alignment = data_align
        ws1.cell(row, base_col + 2).border = thin_border

        ws1.cell(row, base_col + 3, r.get('竞争指数', '')).font = data_font
        ws1.cell(row, base_col + 3).alignment = data_align
        ws1.cell(row, base_col + 3).border = thin_border

        ws1.cell(row, base_col + 4, r.get('月搜索指数', '')).font = data_font
        ws1.cell(row, base_col + 4).alignment = data_align
        ws1.cell(row, base_col + 4).border = thin_border

        ws1.cell(row, base_col + 5, r.get('市场出价', '')).font = data_font
        ws1.cell(row, base_col + 5).alignment = data_align
        ws1.cell(row, base_col + 5).border = thin_border

# 品牌对标区（第4组）：提取品牌名作为品牌列
brands = set()
for r in rows:
    kw = r.get('\ufeff关键词', r.get('关键词', ''))
    # 从关键词提取品牌
    for b in ['muji', '无印良品', '宜家', '野兽派', '祖马龙', '祖玛珑',
              'diptyque', '蒂普提克', '观夏', 'dimoo', '泡泡玛特', 'lelabo',
              '名创优品', 'miniso', 'oce', '山姆', 'kkv', 'Trudon',
              '乐欧', 'loewe', '梅森马吉拉', '马吉拉', 'santal',
              '伊索', '欧珑', '茉莉奶白', '毛戈平', '兰蔻', '迪奥',
              'synesmoon', '宋朝', '海螺', '星星人', 'hana']:
        if b.lower() in kw.lower() or b in kw:
            brands.add(b)

brands = sorted(brands)[:10]  # 最多10个品牌列
print(f"品牌对标列: {len(brands)}个品牌")

brand_font = Font(name='Arial', size=9)
for b_idx, bname in enumerate(brands):
    col = 16 + b_idx
    ws1.cell(3, col, bname).fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    ws1.cell(3, col).font = Font(name='Arial', size=10, color='FFFFFF')
    ws1.cell(3, col).alignment = Alignment(horizontal='center', vertical='center')
    ws1.cell(3, col).border = thin_border
    ws1.column_dimensions[get_column_letter(col)].width = 14

    # 标记：哪些行提到这个词
    for i, r in enumerate(rows):
        kw = r.get('\ufeff关键词', r.get('关键词', ''))
        row = 4 + i
        if bname.lower() in kw.lower() or bname in kw:
            ws1.cell(row, col).value = '\u2713'
            ws1.cell(row, col).font = brand_font
            ws1.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        ws1.cell(row, col).border = thin_border

# 自动筛选扩大到品牌对标区
max_col = get_column_letter(max(16 + len(brands) - 1, 15))
ws1.auto_filter.ref = f"A3:{max_col}3"

wb.save(DST)
print(f"v2 saved: {DST}")
