# -*- coding: utf-8 -*-
import shutil, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

src = r'C:\Users\胡康杰\Desktop\竞品调研框架-v15.xlsx'
out = r'C:\Users\胡康杰\Desktop\竞品调研框架-v16.xlsx'

shutil.copy2(src, out)
wb = openpyxl.load_workbook(out)
ws3 = wb['竞品产品对比']
ws2 = wb['小红书热搜产品']

img_dir = r'C:\Users\胡康杰\Desktop\竞品调研脚本\product_imgs'

# ===== 清空所有商品列（B~O）的数据 =====
for r in range(1, 29):
    for c in range(2, 16):
        ws3.cell(r, c).value = None

# 但保留R1的商品链接提示文本
ws3.cell(1, 1).value = '商品链接 >>>'

# ===== 从Sheet2提取14个核心商品 =====
# 按原始Sheet4的商品列顺序，从Sheet2找到对应的品牌
# 14个商品列：B~O

# 我先用Sheet2里有真实淘宝链接的商品，建立映射
# 黑白调系列、爱果乐系列、原始其它品牌
sheet2_data = {}
for r in range(3, ws2.max_row+1):
    name = (ws2.cell(r, 2).value or '') + (ws2.cell(r, 4).value or '')
    link = ws2.cell(r, 6).value or ''
    price = ws2.cell(r, 8).value or ''
    sheet2_data[name] = {'link': link, 'price': price, 'row': r}

# 14个核心商品（尽量选有真实天猫链接的）
# 黑白调s2(id=726606125346) -> 有真实链接 ✅
# 黑白调c2(id=576762938946) -> 有真实链接 ✅
# 黑白调A2(id=696670340098/id=834193648278) -> 有真实链接 ✅
# 爱果乐沉浸岛(id=911278945859) -> 有真实链接 ✅
# 宜家(ikea.cn) -> 有真实链接 ✅
# 爱果乐咖啡猫(id=564710501147) -> 原始有链接
# 尊眠(id=989595995854) -> 原始有链接
# 护童星辰1号(id=854683228413) -> 原始有链接
# 青节(id=985858990364等) -> 原始有链接
# 龙承匠人 -> 原始有链接

# 商品列映射: col -> (品牌, 产品名, sheet2_name)
columns = {
    'B': ('尊眠', '无名字学习桌', '尊眠'),
    'C': ('黑白调', '黑白调s2儿童学习桌', '黑白调s2儿童学习桌'),
    'D': ('黑白调', '黑白调c2儿童学习桌', '黑白调c2儿童学习桌'),
    'E': ('黑白调', '黑白调A2儿童学习桌', '黑白调A2儿童学习桌'),
    'F': ('爱果乐', '爱果乐沉浸岛学习桌', '爱果乐沉浸岛学习桌'),
    'G': ('爱果乐', '爱果乐学习桌咖啡猫', '爱果乐咖啡猫'),
    'H': ('爱果乐', '爱果乐木木岛', '爱果乐木木岛'),
    'I': ('爱果乐', '爱果乐艺简', '爱果乐艺简'),
    'J': ('护童', '护童学习桌', '护童学习桌'),
    'K': ('青节', '青节学习桌', '青节学习桌'),
    'L': ('宜家', '宜家学习桌', '宜家学习桌'),
    'M': ('龙承匠人', '龙承匠人学习桌', '龙承匠人学习桌怎么样'),
    'N': ('枝乐', '枝乐学习桌', '枝乐学习桌'),
    'O': ('源氏木语', '源氏木语学习桌', '源氏木语学习桌'),
}

# ===== 从Sheet2填已有数据 =====
for col_letter, (brand, prod_name, sheet2_key) in columns.items():
    col_idx = ord(col_letter) - ord('A') + 1
    
    # R1: 商品链接
    if sheet2_key in sheet2_data:
        ws3.cell(1, col_idx).value = sheet2_data[sheet2_key]['link']
    
    # R2: 品牌
    ws3.cell(2, col_idx).value = brand
    
    # R3: 产品名
    ws3.cell(3, col_idx).value = prod_name
    
    # R5: 价位
    if sheet2_key in sheet2_data and sheet2_data[sheet2_key]['price']:
        ws3.cell(5, col_idx).value = sheet2_data[sheet2_key]['price']

# ===== 嵌入图片 =====
# 图片名到产品的映射
img_map = {
    '尊眠': None,
    '黑白调s2儿童学习桌': '黑白调s2儿童学习桌',
    '黑白调c2儿童学习桌': '黑白调c2儿童学习桌',
    '黑白调A2儿童学习桌': '黑白调A2儿童学习桌',
    '爱果乐沉浸岛学习桌': '爱果乐沉浸岛学习桌',
    '爱果乐咖啡猫': '爱果乐学习桌咖啡猫',
    '爱果乐木木岛': '爱果乐木木岛',
    '爱果乐艺简': '爱果乐艺简',
    '护童学习桌': '护童学习桌',
    '青节学习桌': '青节学习桌',
    '宜家学习桌': '宜家学习桌',
    '龙承匠人学习桌': '龙承匠人学习桌',
    '枝乐学习桌': '枝乐学习桌',
    '源氏木语学习桌': '源氏木语学习桌',
}

for col_letter, (brand, prod_name, sheet2_key) in columns.items():
    col_idx = ord(col_letter) - ord('A') + 1
    img_key = img_map.get(prod_name)
    if img_key:
        img_path = None
        for f in os.listdir(img_dir):
            if f.endswith('.jpg') and os.path.getsize(os.path.join(img_dir,f)) > 2000:
                base = f.replace('.jpg', '')
                if base == img_key or base.startswith(img_key):
                    img_path = os.path.join(img_dir, f)
                    break
        if img_path:
            try:
                img = XlImage(img_path)
                img.width = 100
                img.height = 100
                cell_ref = f'{col_letter}4'
                img.anchor = cell_ref
                ws3.add_image(img)
                print(f'  ✅ {col_letter}4: {prod_name}')
            except Exception as e:
                print(f'  ❌ {col_letter}: {e}')

# ===== 恢复样式 =====
param_font = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
param_fill = PatternFill(start_color='6699CC', end_color='6699CC', fill_type='solid')
param_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
data_font = Font(name='微软雅黑', size=9)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
link_font = Font(name='微软雅黑', size=8, color='2563EB')
link_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
brand_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# A列样式
for r in range(1, 29):
    cell = ws3.cell(r, 1)
    cell.font = param_font
    cell.fill = param_fill
    cell.alignment = param_align
    cell.border = thin_border

# 商品列样式
for r in range(1, 29):
    for c in range(2, 16):
        cell = ws3.cell(r, c)
        cell.font = data_font
        cell.alignment = data_align
        cell.border = thin_border
        if r == 1:
            cell.font = link_font
            cell.alignment = link_align
        if r == 2:
            cell.fill = brand_fill

# 行高/列宽（同原始）
row_heights = {
    1: 67.5, 2: 14.25, 3: 42.75, 4: 42.5,
    5: 28.5, 6: 28.5, 7: 28.5, 8: 28.5,
    9: 42.75, 10: 28.5, 11: 14.25, 12: 28.5,
    13: 14.25, 14: 28.5, 15: 14.25, 16: 14.25,
    17: 42.75, 18: 28.5, 19: 28.5, 20: 28.5,
    21: 14.25, 22: 14.25, 23: 14.25, 24: 28.5,
    25: 14.25, 26: 14.25, 27: 14.25, 28: 142.5,
}
col_widths = {
    'A': 11.6, 'B': 13.0, 'C': 13.0, 'D': 11.25, 'E': 13.0,
    'F': 13.0, 'G': 13.0, 'H': 13.0, 'I': 13.0, 'J': 14.9,
    'K': 13.0, 'L': 13.0, 'M': 13.0, 'N': 13.0, 'O': 19.0,
}

for r, h in row_heights.items():
    ws3.row_dimensions[r].height = h
for col_l, w in col_widths.items():
    ws3.column_dimensions[col_l].width = w

wb.save(out)
wb.close()
print(f'\n✅ 保存至: {out}')
