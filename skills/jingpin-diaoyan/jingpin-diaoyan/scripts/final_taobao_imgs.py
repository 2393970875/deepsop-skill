# -*- coding: utf-8 -*-
"""最终版：嵌入所有淘宝真实主图到rebuild文件"""
import openpyxl, os
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

DESKTOP = r"C:\Users\胡康杰\Desktop"
DST = os.path.join(DESKTOP, "竞品调研框架-香薰蜡烛-rebuild.xlsx")
IMG_DIR = os.path.join(DESKTOP, "candle_taobao_imgs")

wb = openpyxl.load_workbook(DST)

thin = Border(left=Side(style='thin',color='D1D5DB'),right=Side(style='thin',color='D1D5DB'),top=Side(style='thin',color='D1D5DB'),bottom=Side(style='thin',color='D1D5DB'))

# 检查有哪些jpeg
img_map = {}
for f in os.listdir(IMG_DIR):
    if f.endswith('.jpeg'):
        brand = f.replace('.jpeg', '').replace('_src', '').replace('taobao', '')
        img_map[brand] = os.path.join(IMG_DIR, f)
        print(f"  图片: {f} -> {brand}")

# 品牌映射：Sheet2行->品牌关键词
s2_brands = {
    2: 'diptyque', 3: 'diptyque',
    4: 'muji', 5: 'muji',
    6: '祖玛珑', 7: '祖玛珑',
    8: '野兽派2', 9: '野兽派2',
    10: '观夏',
    11: '名创优品', 12: '名创优品',
    13: '宜家',
    14: 'lelabo', 15: 'lelabo',
}

ws2 = wb['小红书热搜产品']

for row, brand in s2_brands.items():
    matched = None
    for k, v in img_map.items():
        if brand.lower() in k.lower() or k.lower() in brand.lower():
            matched = v
            break
    if matched:
        try:
            xl = XlImage(matched)
            xl.width = 130; xl.height = 130
            xl.anchor = f'F{row}'
            ws2.add_image(xl)
            ws2.row_dimensions[row].height = 145
            print(f"Sheet2 F{row} ({brand}): ✅")
        except Exception as e:
            print(f"Sheet2 F{row} ({brand}): ❌ {e}")

# Sheet3：列->品牌
s3_brands = {2:'diptyque',3:'muji',4:'祖玛珑',5:'野兽派2',
             6:'观夏',7:'名创优品',8:'宜家',9:'lelabo'}

ws3 = wb['竞品产品对比']
for col, brand in s3_brands.items():
    matched = None
    for k, v in img_map.items():
        if brand.lower() in k.lower() or k.lower() in brand.lower():
            matched = v
            break
    if matched:
        try:
            xl = XlImage(matched)
            xl.width = 100; xl.height = 100
            xl.anchor = f'{get_column_letter(col)}4'
            ws3.add_image(xl)
            print(f"Sheet3 {get_column_letter(col)}4 ({brand}): ✅")
        except Exception as e:
            print(f"Sheet3 {get_column_letter(col)}4 ({brand}): ❌ {e}")

# 补Sheet3链接
links = {
    2: "https://detail.tmall.com/item.htm?id=623852508377",
    3: "https://www.muji.com.cn/",
    4: "https://www.jomalone.com.cn/",
    5: "https://detail.tmall.com/item.htm?id=639150643595",
    6: "https://detail.tmall.com/item.htm?id=727529044639",
    7: "https://www.miniso.cn/",
    8: "https://www.ikea.cn/collections/scents-cleaning",
    9: "https://www.lelabofragrances.com/",
}
for col, url in links.items():
    cell = ws3.cell(1, col, url)
    cell.font = Font(name='微软雅黑', size=8, color='2563EB', underline='single')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin

wb.save(DST)
print(f"\n✅ 最终版: {DST}")
