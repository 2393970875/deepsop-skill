# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r'C:\Users\胡康杰\Desktop\竞品调研框架-v22.xlsx')
ws3 = wb['竞品产品对比']
ws2 = wb['小红书热搜产品']

# ===== 从Sheet2提取所有商品的链接 =====
sheet2_links = {}
for r in range(3, ws2.max_row + 1):
    # B列品牌, D列产品, F列链接
    brand = ws2.cell(r, 2).value or ''
    prod = ws2.cell(r, 4).value or ''
    link = ws2.cell(r, 6).value or ''
    if link and not str(link).startswith('=_xlfn.DISPIMG') and str(link).startswith('http'):
        key = brand + prod
        sheet2_links[key] = link
        # 也用单独的品牌/产品名匹配
        if prod:
            sheet2_links[prod] = link
        if brand:
            sheet2_links[brand] = link

print('Sheet2可用链接:')
for k, v in sheet2_links.items():
    print(f'  {k}: {str(v)[:50]}')

# ===== 填充Sheet3缺的链接 =====
for c in range(2, 16):
    cl = get_column_letter(c)
    existing_link = ws3.cell(1, c).value
    brand = str(ws3.cell(2, c).value or '')
    prod = str(ws3.cell(3, c).value or '')
    
    if not existing_link or not str(existing_link).startswith('http'):
        # 从Sheet2找匹配
        found = None
        # 先精确匹配产品名
        if prod in sheet2_links:
            found = sheet2_links[prod]
        # 再品牌+产品关键词
        elif brand in sheet2_links:
            found = sheet2_links[brand]
        else:
            # 模糊匹配
            for key, link in sheet2_links.items():
                if prod[:6] in key or key[:6] in prod:
                    found = link
                    break
        
        if found:
            ws3.cell(1, c).value = found
            print(f'  ✅ {cl} ({brand}/{prod}) 链接已补: {str(found)[:50]}')
        else:
            # 用淘宝搜索链接兜底
            search_url = f'https://s.taobao.com/search?q={prod}'
            ws3.cell(1, c).value = search_url
            print(f'  ⚠️ {cl} ({brand}/{prod}) 无真实链接，已填搜索链接')

wb.save(r'C:\Users\胡康杰\Desktop\竞品调研框架-v22.xlsx')
wb.close()

print('\n✅ 完成!')
# 最终检查
wb2 = openpyxl.load_workbook(r'C:\Users\胡康杰\Desktop\竞品调研框架-v22.xlsx')
ws = wb2['竞品产品对比']
for c in range(2, 16):
    cl = get_column_letter(c)
    link = ws.cell(1, c).value
    prod = ws.cell(3, c).value
    status = '✅' if link and str(link).startswith('http') else '❌'
    print(f'  {status} {cl} {prod}: {str(link or "无")[:55]}')
wb2.close()
