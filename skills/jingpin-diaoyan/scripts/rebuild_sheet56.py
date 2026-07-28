# -*- coding: utf-8 -*-
import shutil, copy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 基于当前最新v26
src = r'C:\Users\胡康杰\Desktop\竞品调研框架-v26.xlsx'
out = r'C:\Users\胡康杰\Desktop\竞品调研框架-v27.xlsx'

shutil.copy2(src, out)
wb = openpyxl.load_workbook(out)

# ===== Sheet5: 行业&竞品爆文分析（对标原始8.行业&竞品爆文内容分析-爱果乐）=====
ws5 = wb['行业&竞品爆文分析']

# 清空全部
for r in range(1, ws5.max_row+1):
    for c in range(1, 33):
        ws5.cell(r, c).value = None
        ws5.cell(r, c).font = Font()
        ws5.cell(r, c).fill = PatternFill(fill_type=None)
        ws5.cell(r, c).alignment = Alignment()
        ws5.cell(r, c).border = Border()

for m in list(ws5.merged_cells.ranges):
    ws5.unmerge_cells(str(m))

# 列宽（精确匹配原始）
col_widths_5 = {
    'A': 39.0, 'B': 9.0, 'C': 20.375, 'D': 13.0, 'E': 13.0,
    'F': 9.0, 'G': 18.0, 'H': 12.125, 'I': 5.625, 'J': 13.0,
    'K': 13.0, 'L': 13.0, 'M': 13.0, 'N': 13.0, 'O': 13.0,
    'P': 13.0, 'Q': 13.0, 'R': 13.0, 'S': 13.0, 'T': 13.0,
    'U': 13.0, 'V': 13.0, 'W': 13.0, 'X': 13.0, 'Y': 13.0,
    'Z': 13.0, 'AA': 13.0, 'AB': 13.0, 'AC': 13.0, 'AD': 13.0,
    'AE': 13.0, 'AF': 13.0,
}
for cl, w in col_widths_5.items():
    ws5.column_dimensions[cl].width = w

# 行高（原始R1=49.5，后续看需要留50行空行）
row_heights_5 = {1: 49.5}
for r in range(2, 51):
    row_heights_5[r] = 49.5

for r, h in row_heights_5.items():
    ws5.row_dimensions[r].height = h

# R1 字段行（32列）
fields_5 = [
    '笔记归类', '笔记发布时间', '笔记标题', '笔记封面', '笔记类型', '笔记形式',
    '报备合作品牌', '种草提及品牌', '是否品牌合作人', '是否商业笔记', '是否推广笔记',
    '外溢数据', '', '笔记类型', '笔记链接', '互动量', '点赞', '收藏', '评论', '分享',
    '达人昵称', '达人ID', '粉丝数', '达人属性', '地域', '达人小红书主页地址',
    '达人千瓜主页地址', '邮箱', '笔记涵盖热搜词Top10', '笔记相关评论热词Top10',
    '达人标签(前5)', '笔记内容特征'
]

# R1 样式：原始看起来是白底黑字11号宋体/微软雅黑
h1_font = Font(name='微软雅黑', size=9, bold=False)
h1_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
h1_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

for i, f in enumerate(fields_5):
    c = i + 1
    cell = ws5.cell(1, c)
    cell.value = f if f else None
    cell.font = h1_font
    cell.alignment = h1_align
    cell.fill = h1_fill
    cell.border = thin_border

# C13（笔记类型）和C14之间有空一列（C13=空）
# 原始中C13是空的，C14是"笔记类型"（第二次）
# 实际上32列中原始R2: C12=外溢数据(2178), C14=笔记类型(家居用品)
# 所以C13是空的

# R2~R50: 空数据行，保持同样的边框和样式
data_font = Font(name='微软雅黑', size=9)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
for r in range(2, 51):
    for c in range(1, 33):
        cell = ws5.cell(r, c)
        cell.font = data_font
        cell.alignment = data_align
        cell.fill = h1_fill
        cell.border = thin_border
    # 设置行高可自动扩展
    if r not in row_heights_5:
        pass  # 默认

# ===== Sheet6: 爆文拆解（对标原始爆文拆解）=====
ws6 = wb['爆文拆解']

# 清空
for r in range(1, ws6.max_row+1):
    for c in range(1, 20):
        ws6.cell(r, c).value = None
        ws6.cell(r, c).font = Font()
        ws6.cell(r, c).fill = PatternFill(fill_type=None)
        ws6.cell(r, c).alignment = Alignment()
        ws6.cell(r, c).border = Border()

for m in list(ws6.merged_cells.ranges):
    ws6.unmerge_cells(str(m))

# 列宽（精确匹配原始）
col_widths_6 = {
    'A': 9.375, 'B': 9.0, 'C': 5.375, 'D': 4.625, 'E': 13.0,
    'F': 17.125, 'G': 5.0, 'H': 16.875, 'I': 52.5, 'J': 52.5,
    'K': 13.0, 'L': 13.0, 'M': 13.0, 'N': 13.0, 'O': 13.0,
    'P': 13.0, 'Q': 13.0, 'R': 13.0, 'S': 13.0,
}
for cl, w in col_widths_6.items():
    ws6.column_dimensions[cl].width = w

# 行高
row_heights_6 = {1: 34.5, 2: 18.75, 3: 51.75, 4: 172.5, 5: 190.5, 6: 295.5, 7: 222.75, 8: 192.0}
for r in range(1, 31):
    ws6.row_dimensions[r].height = row_heights_6.get(r, 120)

# 合并单元格
ws6.merge_cells('B1:I1')    # 复制粘贴（横跨B~I）
ws6.merge_cells('J1:N1')    # 归纳盘点，与后续建立内容库对应
ws6.merge_cells('A2:A3')    # 更新日期
ws6.merge_cells('B2:E2')    # 笔记情况
ws6.merge_cells('F2:I2')    # 底层拆解
ws6.merge_cells('J2:P2')    # 笔记分析

# 颜色
white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
orange_fill = PatternFill(start_color='FFFFC60A', end_color='FFFFC60A', fill_type='solid')
blue_fill = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
base_font = Font(name='微软雅黑', size=9, bold=False, color='FF1F2329')
bold_font = Font(name='微软雅黑', size=9, bold=True, color='FF1F2329')
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# R1
ws6.cell(1, 2).value = '复制粘贴'
ws6.cell(1, 2).font = base_font
ws6.cell(1, 2).fill = white_fill
ws6.cell(1, 2).alignment = center_align
ws6.cell(1, 2).border = thin_border

ws6.cell(1, 10).value = '归纳盘点，与后续建立内容库对应'
ws6.cell(1, 10).font = base_font
ws6.cell(1, 10).fill = white_fill
ws6.cell(1, 10).alignment = center_align
ws6.cell(1, 10).border = thin_border

ws6.cell(1, 16).value = '梳理选题/封面/标题/文案中好的部分\n归纳其中易传播、可复用的'
ws6.cell(1, 16).font = base_font
ws6.cell(1, 16).fill = white_fill
ws6.cell(1, 16).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws6.cell(1, 16).border = thin_border

# R2
r2_font = Font(name='微软雅黑', size=9, bold=False, color='FF1F2329')

ws6.cell(2, 1).value = '更新日期'
ws6.cell(2, 1).font = r2_font
ws6.cell(2, 1).fill = orange_fill
ws6.cell(2, 1).alignment = center_align
ws6.cell(2, 1).border = thin_border

ws6.cell(2, 2).value = '笔记情况'
ws6.cell(2, 2).font = r2_font
ws6.cell(2, 2).fill = blue_fill
ws6.cell(2, 2).alignment = center_align
ws6.cell(2, 2).border = thin_border

ws6.cell(2, 6).value = '底层拆解'
ws6.cell(2, 6).font = r2_font
ws6.cell(2, 6).fill = blue_fill
ws6.cell(2, 6).alignment = center_align
ws6.cell(2, 6).border = thin_border

ws6.cell(2, 10).value = '笔记分析'
ws6.cell(2, 10).font = r2_font
ws6.cell(2, 10).fill = blue_fill
ws6.cell(2, 10).alignment = center_align
ws6.cell(2, 10).border = thin_border

# R3 子字段
fields_6 = ['', '链接', '点赞', '收藏', '评论', '封面', '笔记类型', '标题', '文案+标签',
            '内页附图/视频结构', '选题风格', '封面形式', '标题结构', '文案结构', '评论区',
            '整体亮点', '全部数据', '免费流量', '外溢进店量']

for i, f in enumerate(fields_6):
    c = i + 1
    cell = ws6.cell(3, c)
    cell.value = f if f else None
    cell.font = r2_font if f else base_font
    cell.fill = blue_fill if c >= 2 and c <= 5 else white_fill
    if c >= 2 and c <= 5:
        cell.fill = blue_fill
    cell.alignment = center_align
    cell.border = thin_border

# R4~R30: 空数据行
for r in range(4, 31):
    for c in range(1, 20):
        cell = ws6.cell(r, c)
        cell.font = Font(name='微软雅黑', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = white_fill
        cell.border = thin_border

# R2合并区域周边单元格也要有样式（合并区域中非顶格单元格不设置值）
# 设置B2~E2, F2~I2, J2~P2 其他合并区域的边框

wb.save(out)
wb.close()

print(f'✅ 保存至: {out}')

# 验证
wb2 = openpyxl.load_workbook(out)
ws5 = wb2['行业&竞品爆文分析']
ws6 = wb2['爆文拆解']
print(f'\n行业&竞品爆文分析: {ws5.max_row}行 x {ws5.max_column}列')
print(f'  A列宽: {ws5.column_dimensions["A"].width}')
print(f'  R1C1: {ws5.cell(1,1).value}')

print(f'\n爆文拆解: {ws6.max_row}行 x {ws6.max_column}列')
print(f'  A列宽: {ws6.column_dimensions["A"].width}')
print(f'  合并: {list(ws6.merged_cells.ranges)}')
print(f'  R2C1 fill: {ws6.cell(2,1).fill.start_color.rgb}')
print(f'  R2C2 fill: {ws6.cell(2,2).fill.start_color.rgb}')
wb2.close()
