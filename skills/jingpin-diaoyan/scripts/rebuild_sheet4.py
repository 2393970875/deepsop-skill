# -*- coding: utf-8 -*-
import shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

src = r'C:\Users\胡康杰\Desktop\竞品调研框架-v23.xlsx'
out = r'C:\Users\胡康杰\Desktop\竞品调研框架-v25.xlsx'
orig = r'C:\Users\胡康杰\Desktop\0613-竞品调研.xlsx'

shutil.copy2(src, out)
wb = openpyxl.load_workbook(out)
ws = wb['关键词排名现状']
orig_wb = openpyxl.load_workbook(orig)
orig_ws = orig_wb['5.关键词排名现状']

# ===== 1. 先取消所有合并 =====
for m in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(m))

# ===== 2. 清空 =====
for r in range(1, 15):
    for c in range(1, 23):
        cell = ws.cell(r, c)
        cell.value = None
        cell.font = Font()
        cell.fill = PatternFill(fill_type=None)
        cell.alignment = Alignment()
        cell.border = Border()

# ===== 3. 列宽（精确匹配原始） =====
col_widths = {
    'A': 9.0, 'B': 6.5, 'C': 7.375, 'D': 13.0, 'E': 13.0,
    'F': 12.625, 'G': 14.75, 'H': 12.6083333333333, 'I': 7.375,
    'J': 13.0, 'K': 13.0, 'L': 11.25, 'M': 15.1916666666667,
    'N': 9.0, 'O': 13.0, 'P': 20.5, 'Q': 69.3083333333333,
    'R': 7.9, 'S': 13.0, 'T': 12.875, 'U': 38.1916666666667,
    'V': 11.025,
}
for cl, w in col_widths.items():
    ws.column_dimensions[cl].width = w

# ===== 4. 行高 =====
row_heights = {
    1: None, 2: None, 3: None, 4: 33.75,
    5: 100.5, 6: 100.5, 7: 100.5, 8: 100.5,
    9: None, 10: 179.15, 11: 173.35, 12: None, 13: None, 14: None,
}
for r, h in row_heights.items():
    ws.row_dimensions[r].height = h if h else None

# ===== 5. 合并单元格 =====
ws.merge_cells('B3:K3')   # 近90天品牌笔记数据
ws.merge_cells('L3:Q3')   # 专业矩阵号
ws.merge_cells('R3:T4')   # 回淘/推广 (R3~T4跨行跨列合并)

# ===== 6. R3 分类行样式 =====
# 原始：微软雅黑11号，居中，非加粗，无特殊填色（theme=1表示默认）
# 但B3的fill实际是FFFFC000
cat_font = Font(name='微软雅黑', size=11)
cat_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cat_fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')

# B3(近90天品牌笔记数据) - 有黄底
ws.cell(3, 2).value = '近90天品牌笔记数据'
ws.cell(3, 2).font = cat_font
ws.cell(3, 2).alignment = cat_align
ws.cell(3, 2).fill = cat_fill

# L3(专业矩阵号) 在原始fill是FFFFC000还是空？从之前输出看到填色=FFFFC000
ws.cell(3, 12).value = '专业矩阵号'
ws.cell(3, 12).font = cat_font
ws.cell(3, 12).alignment = cat_align
ws.cell(3, 12).fill = cat_fill

# R3(回淘/推广)被R3:T4合并覆盖，写入R3即可
ws.cell(3, 18).value = '回淘/推广'
ws.cell(3, 18).font = cat_font
ws.cell(3, 18).alignment = cat_align
ws.cell(3, 18).fill = cat_fill

# ===== 7. R4 字段行 =====
# 原始R4所有单元格：FFFFC000黄底，微软雅黑11号，居中，自动换行
# A4(品牌)非加粗，B4起大部分加粗
# 从之前输出看 R4C1 fill=None(00000000)/字体非粗体; R4C2 fill=FFFFC000/粗体
# 原始中所有R4字段都是FFFFC000黄底
r4_fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
r4_font_normal = Font(name='微软雅黑', size=11, bold=False)
r4_font_bold = Font(name='微软雅黑', size=11, bold=True)
r4_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

fields = {
    1: ('品牌', False),
    2: ('品牌笔记数量', True),
    3: ('商业笔记数量', True),
    4: ('预估投放金额', True),
    5: ('笔记爆文率\n（互动赞、评、藏＞1000)', True),
    6: ('图文笔记', True),
    7: ('视频笔记', True),
    8: ('笔记预估阅读', True),
    9: ('笔记平均阅读', True),
    10: ('笔记总互动数', True),
    11: ('笔记平均互动', True),
    12: ('矩阵号', False),
    13: ('主页', True),
    14: ('店铺', True),
    15: ('直播', True),
    16: ('闭环电商', True),
    17: ('总结', True),
    # 18~20被R3:T4合并覆盖，不写
    21: ('儿童家具小红书闭环电商排名', False),
    22: ('', False),
}

for c, (v, bold) in fields.items():
    cell = ws.cell(4, c)
    cell.value = v
    cell.font = r4_font_bold if bold else r4_font_normal
    cell.alignment = r4_align
    cell.fill = r4_fill

# ===== 8. R5~R11 数据行（空框架） =====
# 原始：R5~R8是品牌数据行，R9空，R10~R11是总结
# R5~R8样式：A列左对齐粗体微软雅黑10号，数据行居中，C17左边自动换行
data_font = Font(name='微软雅黑', size=10)
data_bold = Font(name='微软雅黑', size=10, bold=True)
data_align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
no_fill = PatternFill(fill_type=None)

for r in [5, 6, 7, 8]:
    for c in range(1, 23):
        cell = ws.cell(r, c)
        cell.font = data_font
        if c == 1:
            cell.font = data_bold
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        elif c == 17:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        else:
            cell.alignment = data_align_center

# R9 留空行
# R10~R11 总结行 
for r in [10, 11]:
    for c in range(1, 23):
        cell = ws.cell(r, c)
        cell.font = data_font
        cell.alignment = data_align_left

# ===== 9. 边框 =====
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

for r in range(4, 12):
    for c in range(1, 23):
        try:
            ws.cell(r, c).border = thin_border
        except:
            pass  # MergedCell不放边框

orig_wb.close()
wb.save(out)
wb.close()

print(f'✅ 保存至: {out}')

# 验证
wb2 = openpyxl.load_workbook(out)
ws2 = wb2['关键词排名现状']
print(f'行数: {ws2.max_row}, 列数: {ws2.max_column}')
print(f'合并: {list(ws2.merged_cells.ranges)}')
print(f'R3B3: {ws2.cell(3,2).value}/fill={ws2.cell(3,2).fill.start_color.rgb}')
print(f'R4C1: {ws2.cell(4,1).value}/bold={ws2.cell(4,1).font.bold}')
print(f'R4C7: {ws2.cell(4,7).value}')
print(f'R4C17: {ws2.cell(4,17).value}')
print(f'A列宽: {ws2.column_dimensions["A"].width}')
print(f'R4行高: {ws2.row_dimensions[4].height}')
print(f'R5行高: {ws2.row_dimensions[5].height}')
wb2.close()
