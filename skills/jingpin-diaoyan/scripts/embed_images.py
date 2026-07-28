# -*- coding: utf-8 -*-
import shutil, os
import openpyxl
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils import get_column_letter

src = r'C:\Users\胡康杰\Desktop\竞品调研框架-v20.xlsx'  # 回到v20（干净版本）
out = r'C:\Users\胡康杰\Desktop\竞品调研框架-v21.xlsx'

shutil.copy2(src, out)
wb = openpyxl.load_workbook(out)
ws3 = wb['竞品产品对比']
ws2 = wb['小红书热搜产品']

img_dir = r'C:\Users\胡康杰\Desktop\竞品调研脚本\product_imgs'

# ===== Sheet3：清空所有图片重新嵌入 =====
# 清掉所有旧图片
ws3._images = []

# 商品到图片的映射
img_map = {
    '无名字学习桌': None,
    '黑白调s2儿童学习桌': '黑白调s2儿童学习桌',
    '黑白调c2儿童学习桌': '黑白调c2儿童学习桌',
    '黑白调A2儿童学习桌': '黑白调A2儿童学习桌',
    '爱果乐沉浸岛学习桌': '爱果乐沉浸岛学习桌',
    '爱果乐学习桌咖啡猫': '爱果乐学习桌咖啡猫',
    '爱果乐木木岛': '爱果乐木木岛',
    '爱果乐艺简': '爱果乐艺简',
    '护童学习桌': '护童学习桌',
    '青节学习桌': '青节学习桌',
    '宜家学习桌': '宜家学习桌',
    '龙承匠人学习桌': '龙承匠人学习桌',
    '枝乐学习桌': '枝乐学习桌',
    '源氏木语学习桌': '源氏木语学习桌',
}

for c in range(2, 16):
    col_l = get_column_letter(c)
    prod = ws3.cell(3, c).value or ''
    img_key = img_map.get(prod)
    if img_key is None:
        continue
    # 找图片文件
    found = None
    for f in os.listdir(img_dir):
        if f.endswith('.jpg') and os.path.getsize(os.path.join(img_dir, f)) > 2000:
            base = f.replace('.jpg', '')
            if base == img_key or base.startswith(img_key):
                found = os.path.join(img_dir, f)
                break
    if found:
        try:
            img = XlImage(found)
            img.width = 130
            img.height = 130
            img.anchor = f'{col_l}4'
            ws3.add_image(img)
            print(f'  ✅ Sheet3 {col_l}4: {prod}')
        except Exception as e:
            print(f'  ❌ Sheet3 {col_l}: {e}')

# ===== Sheet2：调大图片 =====
# 清掉旧的重新嵌
ws2._images = []

# Sheet2的品牌到图片的映射
for r in range(3, ws2.max_row + 1):
    for col in [2, 4]:
        val = ws2.cell(r, col).value
        if val:
            kw = str(val).strip()
            # 从img_map找
            for img_key in img_map.values():
                if img_key and (kw.startswith(img_key[:4]) or img_key.startswith(kw[:4])):
                    found = None
                    for f in os.listdir(img_dir):
                        if f.endswith('.jpg') and os.path.getsize(os.path.join(img_dir,f)) > 2000:
                            base = f.replace('.jpg', '')
                            if base == img_key or base.startswith(img_key):
                                found = os.path.join(img_dir, f)
                                break
                    if found:
                        try:
                            img = XlImage(found)
                            img.width = 130
                            img.height = 130
                            img.anchor = f'G{r}'
                            ws2.add_image(img)
                            print(f'  ✅ Sheet2 G{r}: {kw}')
                        except Exception as e:
                            print(f'  ❌ Sheet2 G{r}: {e}')
                    break
            break

# ===== 行高统一调整 =====
# Sheet3 R4图片行
ws3.row_dimensions[4].height = 130

# Sheet2所有包含图片的行
for r in range(3, ws2.max_row + 1):
    ws2.row_dimensions[r].height = 130
ws2.column_dimensions['G'].width = 18

wb.save(out)
wb.close()

print(f'\n✅ 保存至: {out}')
print(f'Sheet3图片: {len(openpyxl.load_workbook(out)["竞品产品对比"]._images)} 张')
print(f'Sheet2图片: {len(openpyxl.load_workbook(out)["小红书热搜产品"]._images)} 张')
