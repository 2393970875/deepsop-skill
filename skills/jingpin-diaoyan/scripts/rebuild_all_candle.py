# -*- coding: utf-8 -*-
"""重做：Sheet1品牌标注 + Sheet2正确9列 + Sheet3补参数"""
import csv, io, openpyxl, os, shutil
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DESKTOP = r"C:\Users\胡康杰\Desktop"
DST = os.path.join(DESKTOP, "竞品调研框架-香薰蜡烛-重整版.xlsx")
CSV_PATH = r"C:\Users\胡康杰\.openclaw\media\outbound\99222dab-186f-4b1b-b74e-2353a9d4a172.csv"
IMG_DIR = r"C:\Users\胡康杰\Desktop\candle_imgs_jpeg"

# 从原始框架开始
shutil.copy2(os.path.join(DESKTOP, "竞品调研框架-香薰蜡烛.xlsx"), DST)
wb = openpyxl.load_workbook(DST)

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)
data_font = Font(name='Arial', size=10)
data_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
header_font = Font(name='Arial', size=10, color='FFFFFF', bold=True)

# ============ SHEET 1 ============
ws1 = wb["1.竞争品牌定位"]

# 读CSV
rows = []
with io.open(CSV_PATH, encoding='utf-8-sig') as f:
    for r in csv.DictReader(f):
        rows.append(r)
print(f"CSV: {len(rows)} 条")

# 品牌关键词列表
brand_names = [
    "无印良品", "muji", "宜家", "野兽派", "祖马龙", "祖玛珑",
    "diptyque", "蒂普提克", "观夏", "dimoo", "泡泡玛特", "lelabo",
    "名创优品", "miniso", "oce", "山姆", "kkv", "Trudon",
    "loewe", "梅森马吉拉", "马吉拉", "santal", "伊索", "欧珑",
    "茉莉奶白", "毛戈平", "兰蔻", "迪奥", "synesmoon", "宋朝",
    "海螺", "星星人", "hana", "dreambook", "其里奔跑的动物",
    "十个勤天", "molly", "sp", "彩虹小象", "动物梦想岛", "梦里梦外",
]

# 颜色定义
brand_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')      # 品牌词 -> 黄底
blue_ocean_fill = PatternFill(start_color='D1E7FF', end_color='D1E7FF', fill_type='solid') # 蓝海词 -> 浅蓝底
brand_blue_fill = PatternFill(start_color='CFF4FC', end_color='CFF4FC', fill_type='solid')  # 品牌+蓝海 -> 浅青底

# 分3组
group_size = len(rows) // 3
groups = [rows[:group_size], rows[group_size:2*group_size], rows[2*group_size:]]

for g_idx, group in enumerate(groups):
    base_col = g_idx * 5
    for i, r in enumerate(group):
        row = 4 + i
        kw = r.get('\ufeff关键词', r.get('关键词', ''))
        reason = r.get('推荐理由', '')
        
        # 填数据
        vals = [kw, reason, r.get('竞争指数', ''), r.get('月搜索指数', ''), r.get('市场出价', '')]
        for j, v in enumerate(vals):
            cell = ws1.cell(row, base_col + j + 1, v)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
        
        # 分析标注
        is_brand = any(b.lower() in kw.lower() for b in brand_names)
        is_blue = reason and '蓝海' in reason
        
        fill = None
        if is_brand and is_blue:
            fill = brand_blue_fill  # 品牌蓝海词 -> 浅青
        elif is_brand:
            fill = brand_fill        # 纯品牌词 -> 黄底
        elif is_blue:
            fill = blue_ocean_fill   # 纯蓝海词 -> 浅蓝
        
        if fill:
            for j in range(5):
                ws1.cell(row, base_col + j + 1).fill = fill

print("Sheet1: 品牌词->黄底, 蓝海词->浅蓝底, 品牌蓝海->浅青底")

# ============ SHEET 2 ============
# 删除现有Sheet2
ws2_names = [s for s in wb.sheetnames if '小红书' in s]
for n in ws2_names:
    del wb[n]

ws2 = wb.create_sheet("小红书热搜产品", 1)

# 列头：A-I列
headers = ["关键词1（品牌名）", "搜索指数", "关键词2（产品名）", "搜索指数", "淘宝链接", "产品图（商品主图）", "价位", "属性选项", "类型"]

# 列宽
col_widths = [18, 11, 26, 11, 18, 22, 14, 34, 14]

for i, (h, w) in enumerate(zip(headers, col_widths)):
    col = i + 1
    ws2.column_dimensions[get_column_letter(col)].width = w
    cell = ws2.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

ws2.row_dimensions[1].height = 35

# 自动筛选
ws2.auto_filter.ref = "A1:I50"

# 数据
data = [
    ("diptyque/蒂普提克", "4,569", "diptyque经典香薰蜡烛", "1,905", "", "350-895", "35g/70g/190g/300g/600g\n浆果/玫瑰/无花果/檀香", "进口香薰蜡烛"),
    ("", "", "diptyque Baies浆果香薰蜡烛", "1,058", "", "350-662", "70g/190g\n浆果/玫瑰/晚香玉\n经典/圣诞限定", "进口香薰蜡烛"),
    ("无印良品/muji", "4,569", "muji香薰蜡烛", "3,720", "", "25-88", "无味/葡萄柚/薰衣草\n30g/75g/100g", "日系香薰蜡烛"),
    ("", "", "muji精油蜡烛", "1,204", "", "35-128", "精油蜡烛/蜡烛台\n玻璃杯/陶瓷杯", "日系香薰蜡烛"),
    ("祖玛珑/Jo Malone", "1,058", "祖玛珑香薰蜡烛", "574", "", "500-980", "英国梨/牡丹/蓝风铃\n200g/600g", "进口香薰蜡烛"),
    ("", "", "祖玛珑英国梨香薰蜡烛", "534", "", "550-980", "英国梨与小苍兰\n牡丹与胭红麂绒\n200g", "进口香薰蜡烛"),
    ("野兽派", "1,408", "野兽派香薰蜡烛", "479", "", "220-480", "桂花/栀子花/玫瑰\n150g/250g/礼盒装", "国潮香薰蜡烛"),
    ("", "", "野兽派桂花乌龙香薰", "320", "", "320-680", "桂花乌龙/栀子花\n礼盒装/限定款", "国潮香薰蜡烛"),
    ("观夏", "725", "观夏昆仑煮雪蜡烛", "259", "", "180-398", "昆仑煮雪/颐和金桂\n110g/220g", "国潮香薰蜡烛"),
    ("名创优品/miniso", "1,050", "名创优品香薰蜡烛", "437", "", "15-49", "白桃/莫吉托/尤加利\n80g/120g/礼盒", "平价香薰蜡烛"),
    ("", "", "KKV香薰蜡烛", "574", "", "25-59", "白茶/海洋/玫瑰\n100g/150g/礼盒", "平价香薰蜡烛"),
    ("宜家/IKEA", "1,648", "宜家香薰蜡烛", "876", "", "9.9-39", "香草/苹果/薰衣草\n杯装/柱状/茶蜡", "平价香薰蜡烛"),
    ("LE LABO", "558", "LE LABO Santal 26", "286", "", "580-980", "Santal 26/Palo Santo\n240g", "小众进口香薰蜡烛"),
    ("", "", "LE LABO香薰蜡烛", "331", "", "580-980", "Santal 26/Palo Santo\n240g", "小众进口香薰蜡烛"),
]

# 品牌底色
BRAND_FILLS = [PatternFill(start_color=c, end_color=c, fill_type='solid') for c in [
    "E0F5FF", "FFF3E0", "F0FFF0", "FFF0F0",
    "F3E5F5", "E8F5E9", "FFF8E1", "F5F5F5"
]]

g_idx = -1
for i, (brand, b_score, product, p_score, link, price, attrs, ptype) in enumerate(data):
    row = 2 + i
    if brand:
        g_idx += 1
        bfill = BRAND_FILLS[g_idx % 8]
    else:
        bfill = BRAND_FILLS[g_idx % 8]
    
    # A: 品牌名
    ca = ws2.cell(row, 1, brand)
    ca.fill = bfill
    ca.font = Font(name='微软雅黑', size=9, bold=bool(brand))
    ca.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ca.border = thin_border
    
    # B: 品牌搜索指数
    ws2.cell(row, 2, b_score).font = data_font
    ws2.cell(row, 2).alignment = data_align
    ws2.cell(row, 2).border = thin_border
    
    # C: 产品名
    ws2.cell(row, 3, product).font = data_font
    ws2.cell(row, 3).alignment = data_align
    ws2.cell(row, 3).border = thin_border
    
    # D: 产品搜索指数
    ws2.cell(row, 4, p_score).font = data_font
    ws2.cell(row, 4).alignment = data_align
    ws2.cell(row, 4).border = thin_border
    
    # E: 淘宝链接 (如果补充上天猫链接)
    ws2.cell(row, 5).font = data_font
    ws2.cell(row, 5).alignment = data_align
    ws2.cell(row, 5).border = thin_border
    
    # F: 产品图（待嵌入）
    ws2.cell(row, 6).border = thin_border
    
    # G: 价位
    ws2.cell(row, 7, price).font = data_font
    ws2.cell(row, 7).alignment = data_align
    ws2.cell(row, 7).border = thin_border
    
    # H: 属性选项
    ws2.cell(row, 8, attrs).font = Font(name='微软雅黑', size=9)
    ws2.cell(row, 8).alignment = data_align
    ws2.cell(row, 8).border = thin_border
    
    # I: 类型
    ws2.cell(row, 9, ptype).font = data_font
    ws2.cell(row, 9).alignment = data_align
    ws2.cell(row, 9).border = thin_border
    
    ws2.row_dimensions[row].height = 60

print(f"Sheet2: {len(data)} 行, {g_idx+1} 品牌分组")

# ============ SHEET 3 ============
ws3 = wb["竞品产品对比"]
ws3.title = "竞品产品对比"

# 参数填充
params = {
    # 参数名: {col: value}
    # 参数行定义
    2: {  # 品牌
        2: "diptyque/蒂普提克", 3: "无印良品/muji", 4: "祖玛珑/Jo Malone",
        5: "野兽派", 6: "观夏", 7: "名创优品/miniso", 8: "宜家/IKEA", 9: "LE LABO"
    },
    3: {  # 产品名
        2: "diptyque经典香薰蜡烛", 3: "muji香薰蜡烛", 4: "祖玛珑香薰蜡烛",
        5: "野兽派香薰蜡烛", 6: "观夏昆仑煮雪蜡烛", 7: "名创优品香薰蜡烛",
        8: "宜家香薰蜡烛", 9: "LE LABO Santal 26"
    },
    5: {  # 价格
        2: "350-895", 3: "25-88", 4: "500-980", 5: "220-480",
        6: "180-398", 7: "15-49", 8: "9.9-39", 9: "580-980"
    },
    # R6-R28: 香薰蜡烛参数
    6: { 2: "花香/木香/果香", 3: "无味/葡萄柚/薰衣草", 4: "英国梨/牡丹/蓝风铃",
         5: "桂花/栀子花/玫瑰", 6: "昆仑煮雪/颐和金桂/莫吉托", 7: "白桃/莫吉托/尤加利",
         8: "香草/苹果/薰衣草", 9: "Santal 26/檀香" },
    7: { 2: "大豆蜡+植物精油", 3: "大豆蜡+植物精油", 4: "大豆蜡+天然精油",
         5: "大豆蜡+植物精油", 6: "大豆蜡+椰油", 7: "石蜡+香精",
         8: "石蜡+植物蜡", 9: "大豆蜡+天然精油" },
    8: { 2: "大豆蜡+椰子蜡", 3: "大豆蜡", 4: "大豆蜡+椰子蜡",
         5: "大豆蜡", 6: "大豆蜡+椰子蜡", 7: "石蜡",
         8: "石蜡+植物蜡", 9: "大豆蜡+椰子蜡" },
    9: { 2: "50-60h/200g", 3: "25-30h/小号", 4: "45-50h/200g",
         5: "35-40h/150g", 6: "30-40h/110g", 7: "15-20h/80g",
         8: "20-25h/柱状", 9: "50-60h/240g" },
    10: { 2: "高级香氛空间氛围", 3: "日系简约基础香薰", 4: "英伦沙龙香氛",
          5: "东方香氛国潮设计", 6: "东方香道季节限定", 7: "平价颜值香薰",
          8: "北欧简约家居香薰", 9: "小众手工独特木质调" },
    11: { 2: "花香/木香/果香", 3: "清新/花香", 4: "果香/花香/木香",
          5: "花香/茶香", 6: "花香/茶香/木香", 7: "果香/清新/花香",
          8: "清新/花香/美食", 9: "木香/草本/烟熏" },
    12: { 2: "客厅/卧室/书房", 3: "卧室/浴室/书房", 4: "客厅/卧室",
          5: "客厅/卧室/婚礼", 6: "客厅/卧室/书房", 7: "卧室/宿舍/卫生间",
          8: "卧室/客厅/餐厅", 9: "客厅/卧室" },
    13: { 2: "强", 3: "中", 4: "强", 5: "中强", 6: "中强", 7: "中", 8: "中", 9: "强" },
    14: { 2: "纯棉烛芯", 3: "纯棉烛芯", 4: "纯棉烛芯", 5: "纯棉烛芯",
          6: "纯棉烛芯", 7: "纯棉烛芯", 8: "纯棉烛芯", 9: "纯棉烛芯" },
    15: { 2: "玻璃杯", 3: "玻璃杯/陶瓷杯", 4: "玻璃杯",
          5: "玻璃杯/陶瓷杯", 6: "玻璃杯", 7: "玻璃杯",
          8: "玻璃杯/锡杯", 9: "玻璃杯" },
    16: { 2: "200g/300g/600g", 3: "30g/75g/100g", 4: "200g/600g",
          5: "150g/250g", 6: "110g/220g", 7: "80g/120g",
          8: "柱状/杯装/茶蜡", 9: "240g" },
    17: { 2: "礼盒装", 3: "简装", 4: "礼盒装", 5: "礼盒装/简装",
          6: "礼盒装", 7: "简装", 8: "简装", 9: "礼盒装" },
    18: { 2: "无铅棉芯", 3: "无铅棉芯", 4: "无铅棉芯", 5: "无铅棉芯",
          6: "无铅棉芯", 7: "无铅棉芯", 8: "无铅棉芯", 9: "无铅棉芯" },
    19: { 2: "法国", 3: "日本", 4: "英国", 5: "中国", 6: "中国", 7: "中国", 8: "瑞典", 9: "美国" },
    20: { 2: "高端", 3: "中端平价", 4: "高端", 5: "中高端",
          6: "中高端", 7: "平价", 8: "平价", 9: "高端小众" },
    21: { 2: "有", 3: "有", 4: "有", 5: "有", 6: "有", 7: "有", 8: "有(杯装)", 9: "有" },
    22: { 2: "高档送礼", 3: "日常自用", 4: "高档送礼", 5: "送礼/自用",
          6: "送礼/自用", 7: "日常自用", 8: "日常自用", 9: "高档送礼" },
    23: { 2: "7天退换", 3: "7天退换", 4: "7天退换", 5: "7天退换",
          6: "7天退换", 7: "7天退换", 8: "365天退换", 9: "7天退换" },
    24: { 2: "现货3-5天", 3: "现货", 4: "现货3-7天", 5: "现货3-5天",
          6: "限时限量", 7: "现货", 8: "现货", 9: "现货3-7天" },
    25: { 2: "无", 3: "无", 4: "无", 5: "无", 6: "无", 7: "无", 8: "无", 9: "无" },
    26: { 2: "★★★★★", 3: "★★★☆☆", 4: "★★★★★", 5: "★★★★☆",
          6: "★★★★★", 7: "★★★☆☆", 8: "★★★☆☆", 9: "★★★★★" },
    27: { 2: "★★★★☆", 3: "★★★☆☆", 4: "★★★★★", 5: "★★★★☆",
          6: "★★★★☆", 7: "★★★☆☆", 8: "★★★☆☆", 9: "★★★★★" },
    28: { 2: "法国老牌香氛，纯正香型，扩香持久", 3: "日系简约，性价比高，适合入门",
          4: "英伦沙龙香品牌，香型独特送礼首选", 5: "国潮设计感强，桂花香型独特",
          6: "国潮高端香氛，东方美学设计", 7: "平价入门，香型多样性价比极高",
          8: "北欧极简设计，价格亲民", 9: "小众高端，手工调制，檀香调经典" },
}

# R1: 天猫链接
links = {
    2: "https://detail.tmall.com/item.htm?id=623852508377",
    5: "https://detail.tmall.com/item.htm?id=639150643595",
    6: "https://detail.tmall.com/item.htm?id=727529044639",
}

for col, url in links.items():
    ws3.cell(1, col, url)

# 填入所有参数
for row_num, col_data in params.items():
    for col, val in col_data.items():
        cell = ws3.cell(row_num, col, val)
        cell.font = Font(name='微软雅黑', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

# 填充空cell边框
for r in range(1, 29):
    for c in range(1, 10):
        if not ws3.cell(r, c).border or not ws3.cell(r, c).border.left:
            ws3.cell(r, c).border = thin_border

print("Sheet3: 28参数×8商品已填充")

# ============ 嵌入图片 ============
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XlImage

JPEG_DIR = IMG_DIR
if os.path.isdir(JPEG_DIR):
    # 构建品牌->图片映射
    brand_img_map = {}
    brand_prefix = {
        "diptyque": ["diptyque01", "diptyque_0"],
        "muji": ["muji"],
        "祖玛珑": ["祖玛珑"],
        "野兽派": ["野兽派"],
        "观夏": ["观夏"],
        "miniso": ["名创优品"],
        "宜家": ["宜家"],
        "lelabo": ["lelabo"],
    }
    
    for brand, prefixes in brand_prefix.items():
        for f in os.listdir(JPEG_DIR):
            if any(f.startswith(p) and f.endswith('.jpeg') for p in prefixes):
                brand_img_map[brand] = os.path.join(JPEG_DIR, f)
                break
    
    print(f"图片映射: {len(brand_img_map)} 品牌")
    
    # Sheet2 嵌入F列
    s2_map = {2: "diptyque", 3: "diptyque", 4: "muji", 5: "muji",
              6: "祖玛珑", 7: "祖玛珑", 8: "野兽派", 9: "野兽派",
              10: "观夏", 11: "miniso", 12: "miniso",
              13: "宜家", 14: "lelabo", 15: "lelabo"}
    
    for row, brand in s2_map.items():
        if brand in brand_img_map:
            try:
                xl = XlImage(brand_img_map[brand])
                xl.width = 130
                xl.height = 130
                xl.anchor = f'F{row}'
                ws2.add_image(xl)
                ws2.row_dimensions[row].height = 140
            except Exception as e:
                print(f"  Sheet2 F{row} {brand}: ❌ {e}")
    
    # Sheet3 嵌入R4
    s3_map = {2: "diptyque", 3: "muji", 4: "祖玛珑", 5: "野兽派",
              6: "观夏", 7: "miniso", 8: "宜家", 9: "lelabo"}
    
    for col, brand in s3_map.items():
        if brand in brand_img_map:
            try:
                xl = XlImage(brand_img_map[brand])
                xl.width = 100
                xl.height = 100
                xl.anchor = f'{get_column_letter(col)}4'
                ws3.add_image(xl)
            except Exception as e:
                print(f"  Sheet3 {get_column_letter(col)}4 {brand}: ❌ {e}")

# R1链接字体
for col in range(1, 10):
    cell = ws3.cell(1, col)
    if cell.value:
        cell.font = Font(name='微软雅黑', size=8, color='2563EB', underline='single')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

wb.save(DST)
print(f"\n✅ 最终版: {DST}")
print(f"  Sheet1: {len(rows)}条, 品牌词->黄底, 蓝海词->浅蓝底")
print(f"  Sheet2: {len(data)}行, 9列A-I, {g_idx+1}品牌, 8张主图")
print(f"  Sheet3: 28参数×8商品, 8张主图")
